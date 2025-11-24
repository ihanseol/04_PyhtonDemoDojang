"""
Water Quality Data to HWP Document Generator
Extracts water quality data from Excel and generates HWP documents.
"""

import re
import os
import shutil
from pathlib import Path
from enum import Enum
from dataclasses import dataclass
from typing import List, Optional, Tuple

import openpyxl
from pyhwpx import Hwp

from merge_hwp_files import merge_hwp_files


class OpMode(Enum):
    """Operation mode for data processing."""
    SINGLE = 1
    DUAL = 2


@dataclass
class WaterQualityData:
    """Container for water quality measurements."""
    well_name: str
    timestamps: List[str]
    temps: List[float]
    ecs: List[int]
    phs: List[float]

    # For dual mode
    well_name2: Optional[str] = None
    temps2: Optional[List[float]] = None
    ecs2: Optional[List[int]] = None
    phs2: Optional[List[float]] = None

    @property
    def is_dual_mode(self) -> bool:
        """Check if data is in dual mode."""
        return self.well_name2 is not None

    @property
    def temp_stats(self) -> Tuple[float, float]:
        """Get temperature min and max from last 3 values."""
        return self.temps[-2], self.temps[-3]

    @property
    def ec_stats(self) -> Tuple[int, int]:
        """Get EC min and max from last 3 values."""
        return self.ecs[-2], self.ecs[-3]

    @property
    def ph_stats(self) -> Tuple[float, float]:
        """Get pH min and max from last 3 values."""
        return self.phs[-2], self.phs[-3]

    @property
    def temp_stats2(self) -> Optional[Tuple[float, float]]:
        """Get temperature min and max for second well."""
        return (self.temps2[-2], self.temps2[-3]) if self.is_dual_mode else None

    @property
    def ec_stats2(self) -> Optional[Tuple[int, int]]:
        """Get EC min and max for second well."""
        return (self.ecs2[-2], self.ecs2[-3]) if self.is_dual_mode else None

    @property
    def ph_stats2(self) -> Optional[Tuple[float, float]]:
        """Get pH min and max for second well."""
        return (self.phs2[-2], self.phs2[-3]) if self.is_dual_mode else None


class HwpTemplateManager:
    """Manages HWP template files."""

    SINGLE_TEMPLATE = "wt_simple.hwp"
    DUAL_TEMPLATE = "wt_simple_dual.hwp"

    def __init__(self):
        self.desktop = Path(os.environ['USERPROFILE']) / 'Desktop'
        self.template_dir = Path("c:/Program Files/totalcmd/hwp")

    def copy_templates_to_desktop(self) -> bool:
        """Copy both template files to desktop."""
        try:
            shutil.copy(self.template_dir / self.SINGLE_TEMPLATE, self.desktop)
            shutil.copy(self.template_dir / self.DUAL_TEMPLATE, self.desktop)
            return True
        except Exception as e:
            print(f"Error copying templates: {e}")
            return False

    def clean_up(self) -> None:
        """Remove template files from desktop."""
        for template in [self.SINGLE_TEMPLATE, self.DUAL_TEMPLATE]:
            template_path = self.desktop / template
            if template_path.exists():
                template_path.unlink()

    def get_template_path(self, is_dual: bool) -> Path:
        """Get the appropriate template path."""
        template = self.DUAL_TEMPLATE if is_dual else self.SINGLE_TEMPLATE
        return self.desktop / template


class ExcelDataExtractor:
    """Extracts water quality data from Excel files."""

    DATA_START_ROW = 14
    DATA_END_ROW = 24
    STATS_START_ROW = 24
    STATS_END_ROW = 27

    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.workbook = None
        self.last_index = 1

    def load_workbook(self) -> bool:
        """Load the Excel workbook."""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            return True
        except FileNotFoundError:
            print(f"File not found: {self.file_path}")
            return False
        except Exception as e:
            print(f"Error loading workbook: {e}")
            return False

    def extract_all_sheets(self) -> List[WaterQualityData]:
        """Extract data from all sheets in the workbook."""
        if not self.workbook:
            return []

        results = []
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            data = self._extract_sheet_data(sheet)
            if data:
                results.append(data)

        print('self.last_index', self.last_index)
        return results, self.last_index

    def _detect_mode(self, sheet) -> OpMode:
        """Detect if sheet is in single or dual mode."""
        return OpMode.DUAL if sheet['H29'].value is not None else OpMode.SINGLE

    def _extract_sheet_data(self, sheet) -> Optional[WaterQualityData]:
        """Extract water quality data from a single sheet."""
        mode = self._detect_mode(sheet)
        well1 = sheet['D12'].value

        if not well1:
            return None

        print(f"Processing: {well1} ({mode.name} mode)")

        # Initialize data containers
        times, temps1, ecs1, phs1 = [], [], [], []
        temps2, ecs2, phs2 = [], [], []

        # Extract measurement data
        self._extract_measurements(sheet, mode, times, temps1, temps2, ecs1, ecs2, phs1, phs2)

        # Extract statistics
        self._extract_statistics(sheet, mode, temps1, temps2, ecs1, ecs2, phs1, phs2)

        if mode == OpMode.SINGLE:
            self.last_index = self._extract_number(well1)
            return WaterQualityData(well1, times, temps1, ecs1, phs1)
        else:
            well2 = sheet['G12'].value or f"W-{self._extract_number(well1) + 1}"
            self.last_index = self._extract_number(well2)
            return WaterQualityData(well1, times, temps1, ecs1, phs1, well2, temps2, ecs2, phs2)

    def _extract_measurements(self, sheet, mode, times, temps1, temps2, ecs1, ecs2, phs1, phs2):
        """Extract measurement data from rows 14-23."""
        for row in range(self.DATA_START_ROW, self.DATA_END_ROW):
            time_val = sheet.cell(row=row, column=2).value

            # Extract timestamp
            if isinstance(time_val, datetime.datetime):
                fmt = "%y-%#m-%#d %H:%M" if mode == OpMode.DUAL else "%Y-%m-%d %H:%M"
                times.append(time_val.strftime(fmt))

            # Extract measurements for well 1
            self._append_if_numeric(sheet.cell(row=row, column=4).value, temps1, round_to=1)
            self._append_if_numeric(sheet.cell(row=row, column=5).value, ecs1, as_int=True)
            self._append_if_numeric(sheet.cell(row=row, column=6).value, phs1, round_to=2)

            # Extract measurements for well 2 (if dual mode)
            if mode == OpMode.DUAL:
                self._append_if_numeric(sheet.cell(row=row, column=7).value, temps2, round_to=1)
                self._append_if_numeric(sheet.cell(row=row, column=8).value, ecs2, as_int=True)
                self._append_if_numeric(sheet.cell(row=row, column=9).value, phs2, round_to=2)

    def _extract_statistics(self, sheet, mode, temps1, temps2, ecs1, ecs2, phs1, phs2):
        """Extract statistics from rows 24-26."""
        for row in range(self.STATS_START_ROW, self.STATS_END_ROW):
            self._append_if_numeric(sheet.cell(row=row, column=4).value, temps1, round_to=1)
            self._append_if_numeric(sheet.cell(row=row, column=5).value, ecs1, as_int=True)
            self._append_if_numeric(sheet.cell(row=row, column=6).value, phs1, round_to=2)

            if mode == OpMode.DUAL:
                self._append_if_numeric(sheet.cell(row=row, column=7).value, temps2, round_to=1)
                self._append_if_numeric(sheet.cell(row=row, column=8).value, ecs2, as_int=True)
                self._append_if_numeric(sheet.cell(row=row, column=9).value, phs2, round_to=2)

    @staticmethod
    def _append_if_numeric(value, target_list, round_to=None, as_int=False):
        """Append value to list if it's numeric."""
        if isinstance(value, (int, float)):
            if as_int:
                target_list.append(int(value))
            elif round_to is not None:
                target_list.append(round(value, round_to))
            else:
                target_list.append(value)

    @staticmethod
    def _extract_number(text: str) -> int:
        """Extract first number from string."""
        match = re.search(r'\d+', text)
        return int(match.group(0)) if match else 0


class HwpDocumentWriter:
    """Writes water quality data to HWP documents."""

    def __init__(self, template_manager: HwpTemplateManager, output_dir: Path):
        self.template_manager = template_manager
        self.output_dir = output_dir

    def create_document(self, data: WaterQualityData, last_index, i) -> None:
        """Create an HWP document from water quality data."""
        template_path = self.template_manager.get_template_path(data.is_dual_mode)

        hwp = Hwp(visible=False)
        hwp.open(str(template_path))

        try:
            self._write_placeholders(hwp, data, last_index, i)
            self._write_table_data(hwp, data)

            output_path = self.output_dir / f"ex_water_{data.well_name}.hwp"
            hwp.save_as(str(output_path))
            print(f"Created: {output_path}")
        finally:
            hwp.Quit(save=False)

    def _write_placeholders(self, hwp, data: WaterQualityData, last_index, i):
        """Write data to placeholder fields."""
        temp_min, temp_max = data.temp_stats
        ec_min, ec_max = data.ec_stats
        ph_min, ph_max = data.ph_stats

        # 2025/11/25 - add index main value
        self._write_field(hwp, 'index_main', i)

        if data.is_dual_mode:
            self._write_field(hwp, 'index', last_index)

            self._write_field(hwp, 'well1', data.well_name)
            self._write_field(hwp, 'well2', data.well_name2)
            self._write_field(hwp, 'temp1_min', temp_min)
            self._write_field(hwp, 'temp1_max', temp_max)
            self._write_field(hwp, 'ec1_min', ec_min)
            self._write_field(hwp, 'ec1_max', ec_max)
            self._write_field(hwp, 'ph1_min', ph_min)
            self._write_field(hwp, 'ph1_max', ph_max)

            temp2_min, temp2_max = data.temp_stats2
            ec2_min, ec2_max = data.ec_stats2
            ph2_min, ph2_max = data.ph_stats2

            self._write_field(hwp, 'temp2_min', temp2_min)
            self._write_field(hwp, 'temp2_max', temp2_max)
            self._write_field(hwp, 'ec2_min', ec2_min)
            self._write_field(hwp, 'ec2_max', ec2_max)
            self._write_field(hwp, 'ph2_min', ph2_min)
            self._write_field(hwp, 'ph2_max', ph2_max)
        else:
            self._write_field(hwp, 'index', last_index)
            self._write_field(hwp, 'well_main', data.well_name)
            self._write_field(hwp, 'temp_min', temp_min)
            self._write_field(hwp, 'temp_max', temp_max)
            self._write_field(hwp, 'ec_min', ec_min)
            self._write_field(hwp, 'ec_max', ec_max)
            self._write_field(hwp, 'ph_min', ph_min)
            self._write_field(hwp, 'ph_max', ph_max)

    def _write_table_data(self, hwp, data: WaterQualityData):
        """Write data to table cells."""
        table_list = [ctrl for ctrl in hwp.ctrl_list if ctrl.UserDesc == "표"]
        if not table_list:
            return

        hwp.select_ctrl(table_list[0])
        hwp.ShapeObjTableSelCell()

        # Write well names in header
        self._write_cell(hwp, "C2", data.well_name)
        if data.is_dual_mode:
            self._write_cell(hwp, "F2", data.well_name2)

        # Write measurement data (rows 4-13)
        for i in range(4, 14):
            idx = i - 4
            self._write_cell(hwp, f"A{i}", data.timestamps[idx])
            self._write_cell(hwp, f"C{i}", data.temps[idx])
            self._write_cell(hwp, f"D{i}", data.ecs[idx])
            self._write_cell(hwp, f"E{i}", data.phs[idx])

            if data.is_dual_mode:
                self._write_cell(hwp, f"F{i}", data.temps2[idx])
                self._write_cell(hwp, f"G{i}", data.ecs2[idx])
                self._write_cell(hwp, f"H{i}", data.phs2[idx])

        # Write statistics (rows 14-16)
        for i in range(14, 17):
            idx = i - 4
            self._write_cell(hwp, f"C{i}", data.temps[idx])
            self._write_cell(hwp, f"D{i}", data.ecs[idx])
            self._write_cell(hwp, f"E{i}", data.phs[idx])

            if data.is_dual_mode:
                self._write_cell(hwp, f"F{i}", data.temps2[idx])
                self._write_cell(hwp, f"G{i}", data.ecs2[idx])
                self._write_cell(hwp, f"H{i}", data.phs2[idx])

    @staticmethod
    def _write_field(hwp, field_name: str, value):
        """Write value to a field placeholder."""
        hwp.MoveToField(field_name)
        hwp.PutFieldText(field_name, str(value))

    @staticmethod
    def _write_cell(hwp, cell_addr: str, value):
        """Write value to a table cell."""
        hwp.goto_addr(cell_addr)
        hwp.SelectAll()
        hwp.Delete()
        hwp.insert_text(str(value))


def main():
    """Main execution function."""
    excel_file = Path("d:/05_Send/ex_water_test.xlsx")
    output_dir = Path("d:/05_Send")

    if not excel_file.exists():
        print(f"Excel file not found: {excel_file}")
        return

    # Initialize components
    template_mgr = HwpTemplateManager()
    extractor = ExcelDataExtractor(excel_file)
    writer = HwpDocumentWriter(template_mgr, output_dir)

    # Load and extract data
    print("Loading Excel workbook...")
    if not extractor.load_workbook():
        return

    print("Extracting water quality data...")
    all_data, last_index = extractor.extract_all_sheets()

    if not all_data:
        print("No data extracted from Excel file.")
        return

    print(f"Found {len(all_data)} sheets to process\n")

    # Copy templates and create documents
    if not template_mgr.copy_templates_to_desktop():
        print("Failed to copy templates")
        return

    try:
        i = 1
        for data in all_data:
            print(f"\nProcessing {data.well_name}...")
            writer.create_document(data, last_index + i, i)
            i = i + 1

        print("\nMerging HWP files...")
        merge_hwp_files()

        print("\nAll documents created successfully!")
    finally:
        template_mgr.clean_up()


if __name__ == "__main__":
    # Import datetime for type checking
    import datetime

    main()
