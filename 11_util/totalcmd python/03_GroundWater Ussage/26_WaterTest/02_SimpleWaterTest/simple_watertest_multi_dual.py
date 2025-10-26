import re
from pyhwpx import Hwp
import openpyxl
from pathlib import Path
import os
import shutil
import time
import datetime
from merge_hwp_files import merge_hwp_files
from enum import Enum


class OpMode(Enum):
    SINGLE = 1
    DUAL = 2


class GenHwpDoc:
    """Class to generate HWP documents from Excel data."""

    def __init__(self):
        """Initialize the document generator with default paths."""
        self.base_dir = Path("d:/05_Send")
        self.excel_file = "ex_water_test.xlsx"

        self.hwp_template = "wt_simple.hwp"
        self.hwp_template_dual = "wt_simple_dual.hwp"

        self.template_source = Path("c:/Program Files/totalcmd/hwp") / self.hwp_template
        self.template_source2 = Path("c:/Program Files/totalcmd/hwp") / self.hwp_template_dual

        self.desktop = self.get_desktop_path()

    @staticmethod
    def get_desktop_path():
        """Get the path to the user's desktop."""
        return Path(os.environ['USERPROFILE']) / 'Desktop'

    def get_hwp_name(self):
        return self.desktop / self.hwp_template

    def clean_up(self):
        template_path = self.desktop / self.hwp_template
        template_path_dual = self.desktop / self.hwp_template_dual

        if template_path.exists():
            os.remove(template_path)

        if template_path_dual.exists():
            os.remove(template_path_dual)

    def copy_template_to_desktop(self):
        """Copy the HWP template file to the desktop."""
        try:
            # desktop_template = self.desktop / self.hwp_template

            shutil.copy(self.template_source, self.desktop)
            shutil.copy(self.template_source2, self.desktop)
            return True
        except Exception as e:
            print(f"Error copying template to desktop: {str(e)}")
            return False


def analyze_water_quality(file_path):
    opmode = 0

    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)

    except FileNotFoundError:
        print(f"오류 발생: 파일을 찾을 수 없습니다 - {file_path}")
        return None
    except Exception as e:
        print(f"오류 발생: {e}")
        return None

    return_data = []
    sheet_names = workbook.sheetnames

    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]

        chk_cell = sheet['H29'].value

        if chk_cell is None:
            print(f"Cell is empty (Value is None). Single Mode")
            opmode = OpMode.SINGLE
        else:
            print(f"Cell is empty (Value is not None). Dual Mode")
            opmode = OpMode.DUAL

        well1 = sheet['D12'].value
        if opmode == OpMode.DUAL:
            well2 = sheet['G12'].value

        print(f"분석 대상: {well1 if well1 is not None else 'N/A'}")

        # 데이터 리스트 초기화
        times, temps1, temps2, ecs1, ecs2, phs1, phs2 = [], [], [], [], [], [], []

        print('=' * 100)
        for row in range(14, 24):  # 24는 포함되지 않으므로 14행부터 23행까지 반복
            # 셀 값 가져오기
            time_val = sheet.cell(row=row, column=2).value  # B열
            temp1_val = sheet.cell(row=row, column=4).value  # D열
            ec1_val = sheet.cell(row=row, column=5).value  # E열
            ph1_val = sheet.cell(row=row, column=6).value  # F열

            if opmode == OpMode.DUAL:
                temp2_val = sheet.cell(row=row, column=7).value  # G열
                ec2_val = sheet.cell(row=row, column=8).value  # H열
                ph2_val = sheet.cell(row=row, column=9).value  # I열

            # None이 아닌 경우에만 리스트에 추가 (win32com 코드와 동일한 로직)
            # 수치형 데이터인지 확인 (숫자가 아닌 문자열이나 None은 제외)

            if isinstance(time_val, datetime.datetime):
                if opmode == OpMode.SINGLE:
                    times.append(time_val.strftime("%Y-%m-%d %H:%M"))
                    print(time_val.strftime("%Y-%m-%d %H:%M"))
                else:
                    times.append(time_val.strftime("%y-%#m-%#d %H:%M"))
                    print(time_val.strftime("%y-%#m-%#d %H:%M"))

            if isinstance(temp1_val, (int, float)):
                if opmode == OpMode.SINGLE:
                    temps1.append(round(temp1_val, 1))
                else:
                    temps1.append(round(temp1_val, 1))
                    temps2.append(round(temp2_val, 1))

            if isinstance(ec1_val, (int, float)):
                if opmode == OpMode.SINGLE:
                    ecs1.append(int(ec1_val))
                else:
                    ecs1.append(int(ec1_val))
                    ecs2.append(int(ec2_val))

            if isinstance(ph1_val, (int, float)):
                if opmode == OpMode.SINGLE:
                    phs1.append(round(ph1_val, 2))
                else:
                    phs1.append(round(ph1_val, 2))
                    phs2.append(round(ph2_val, 2))

        for row in range(24, 27):  # 27는 포함되지 않으므로 24행부터 26행까지 반복
            # 셀 값 가져오기
            temp1_val = sheet.cell(row=row, column=4).value  # D열
            ec1_val = sheet.cell(row=row, column=5).value  # E열
            ph1_val = sheet.cell(row=row, column=6).value  # F열

            if opmode == OpMode.DUAL:
                temp2_val = sheet.cell(row=row, column=7).value  # G열
                ec2_val = sheet.cell(row=row, column=8).value  # H열
                ph2_val = sheet.cell(row=row, column=9).value  # I열

            if isinstance(temp1_val, (int, float)):
                if opmode == OpMode.SINGLE:
                    temps1.append(round(temp1_val, 1))
                else:
                    temps1.append(round(temp1_val, 1))
                    temps2.append(round(temp2_val, 1))

            if isinstance(ec1_val, (int, float)):
                if opmode == OpMode.SINGLE:
                    ecs1.append(int(ec1_val))
                else:
                    ecs1.append(int(ec1_val))
                    ecs2.append(int(ec2_val))

            if isinstance(ph1_val, (int, float)):
                if opmode == OpMode.SINGLE:
                    phs1.append(round(ph1_val, 2))
                else:
                    phs1.append(round(ph1_val, 2))
                    phs2.append(round(ph2_val, 2))

        if opmode == OpMode.SINGLE:
            results = [well1, times, temps1, ecs1, phs1]
        else:
            results = [well1, times, temps1, temps2, ecs1, ecs2, phs1, phs2]
        return_data.append(results)

    return return_data


def hwp_write_text(hwp, place_holder, data):
    hwp.MoveToField(f"{place_holder}")
    hwp.PutFieldText(f"{place_holder}", f"{data}")


def cell_write(hwp, cell, i, data):
    hwp.goto_addr(f"{cell}{i}")
    hwp.SelectAll()
    hwp.Delete()
    hwp.insert_text(f"{data[i - 4]}")



def hwp_part(results):
    hwp = Hwp(visible=False)

    if len(results) == 5:
        hwp.open(r"c:\Users\minhwasoo\Desktop\wt_simple.hwp")
    else:
        hwp.open(r"c:\Users\minhwasoo\Desktop\wt_simple_dual.hwp")

    if len(results) == 5:
        well1 = results[0]
        times = results[1]
        temps1 = results[2]
        ecs1 = results[3]
        phs1 = results[4]

        ec1_max = ecs1[-3]
        ec1_min = ecs1[-2]

        temp1_max = temps1[-3]
        temp1_min = temps1[-2]

        ph1_max = phs1[-3]
        ph1_min = phs1[-2]
    else:
        well1 = results[0]
        well2 = "W-" + str(extract_number(well1) + 1)
        times = results[1]

        temps1 = results[2]
        temps2 = results[3]

        ecs1 = results[4]
        ecs2 = results[5]

        phs1 = results[6]
        phs2 = results[7]

        ec1_max = ecs1[-3]
        ec1_min = ecs1[-2]
        ec2_max = ecs2[-3]
        ec2_min = ecs2[-2]

        temp1_max = temps1[-3]
        temp1_min = temps1[-2]
        temp2_max = temps2[-3]
        temp2_min = temps2[-2]

        ph1_max = phs1[-3]
        ph1_min = phs1[-2]
        ph2_max = phs2[-3]
        ph2_min = phs2[-2]

    if len(results) == 5:
        hwp_write_text(hwp, 'well_main', well1)
        hwp_write_text(hwp, 'temp_min', temp1_min)
        hwp_write_text(hwp, 'temp_max', temp1_max)

        hwp_write_text(hwp, 'ec_min', ec1_min)
        hwp_write_text(hwp, 'ec_max', ec1_max)

        hwp_write_text(hwp, 'ph_min', ph1_min)
        hwp_write_text(hwp, 'ph_max', ph1_max)
    else:
        hwp_write_text(hwp, 'well1', well1)
        hwp_write_text(hwp, 'well2', well2)

        hwp_write_text(hwp, 'temp1_min', temp1_min)
        hwp_write_text(hwp, 'temp1_max', temp1_max)

        hwp_write_text(hwp, 'temp2_min', temp2_min)
        hwp_write_text(hwp, 'temp2_max', temp2_max)

        hwp_write_text(hwp, 'ec1_min', ec1_min)
        hwp_write_text(hwp, 'ec1_max', ec1_max)

        hwp_write_text(hwp, 'ec2_min', ec2_min)
        hwp_write_text(hwp, 'ec2_max', ec2_max)

        hwp_write_text(hwp, 'ph1_min', ph1_min)
        hwp_write_text(hwp, 'ph1_max', ph1_max)

        hwp_write_text(hwp, 'ph2_min', ph2_min)
        hwp_write_text(hwp, 'ph2_max', ph2_max)

    # 테이블리스트를 가져옴
    Table_list = [i for i in hwp.ctrl_list if i.UserDesc == "표"]
    print(Table_list)

    # 테이블의 첫번째를 선택
    hwp.select_ctrl(Table_list[0])
    hwp.ShapeObjTableSelCell()  # A1 셀을 선택한 상태가 됨

    if len(results) == 5:  # single
        hwp.goto_addr("C2")
        hwp.SelectAll()
        hwp.Delete()
        hwp.insert_text(f"{well1}")

        print(well1)
        print(ph1_max, ph1_min)
        print(temp1_max, temp1_min)
        print(ec1_max, ec1_min)

        for i in range(4, 14):
            cell_write(hwp, "A", i, times)
            cell_write(hwp, "C", i, temps1)
            cell_write(hwp, "D", i, ecs1)
            cell_write(hwp, "E", i, phs1)

        for i in range(14, 17):
            cell_write(hwp, "C", i, temps1)
            cell_write(hwp, "D", i, ecs1)
            cell_write(hwp, "E", i, phs1)
    else:
        hwp.goto_addr("C2")
        hwp.SelectAll()
        hwp.Delete()
        hwp.insert_text(f"{well1}")

        hwp.goto_addr("F2")
        hwp.SelectAll()
        hwp.Delete()
        hwp.insert_text(f"{well2}")

        print(well1, well2)
        print(ph1_max, ph1_min, "//", ph2_max, ph2_min)
        print(temp1_max, temp1_min, "//", temp2_max, temp2_min)
        print(ec1_max, ec1_min, "//", ec2_max, ec2_min)

        for i in range(4, 14):
            cell_write(hwp, "A", i, times)
            cell_write(hwp, "C", i, temps1)
            cell_write(hwp, "D", i, ecs1)
            cell_write(hwp, "E", i, phs1)

            cell_write(hwp, "F", i, temps2)
            cell_write(hwp, "G", i, ecs2)
            cell_write(hwp, "H", i, phs2)

        for i in range(14, 17):
            cell_write(hwp, "C", i, temps1)
            cell_write(hwp, "D", i, ecs1)
            cell_write(hwp, "E", i, phs1)
            # ============================================
            cell_write(hwp, "F", i, temps2)
            cell_write(hwp, "G", i, ecs2)
            cell_write(hwp, "H", i, phs2)

    hwp.save_as(f"d:\\05_Send\\ex_water_{well1}.hwp")
    hwp.Quit(save=False)


def copy_template_to_desktop(self):
    try:
        desktop_template = desktop / hwp_template
        shutil.copy(self.template_source, desktop_template)
        return True
    except Exception as e:
        print(f"Error copying template to desktop: {str(e)}")
        return False


def extract_number(given_string: str) -> int | None:
    """
    Extracts the first continuous sequence of digits from a given string
    and returns it as an integer.

    Args:
        given_string: The string to search (e.g., 'w-20', 'item_34b').

    Returns:
        The extracted number as an integer, or None if no digits are found.
    """
    # Regular expression to find one or more consecutive digits (\d+).
    # re.search scans through the string looking for any location where
    # the pattern produces a match.
    match = re.search(r'\d+', given_string)

    if match:
        # match.group(0) returns the matching substring (e.g., "20")
        # We convert it to an integer before returning.
        return int(match.group(0))

    # Return None if no numeric sequence was found in the string.
    return None


def main():
    excel_file = r"d:\05_Send\ex_water_test.xlsx"
    genhwpdoc = GenHwpDoc()

    wb = openpyxl.load_workbook(excel_file)

    # 모든 시트 이름 가져오기
    sheet_names = wb.sheetnames
    num_sheets = len(sheet_names)

    print(f"발견된 시트: {sheet_names}\n")

    if not os.path.exists(excel_file):
        print(f"파일을 찾을 수 없습니다: {excel_file}")
        print("파일 경로를 확인해주세요. (openpyxl은 .xlsx 파일만 지원합니다.)")
        exit(0)

    print("📢 openpyxl을 사용하여 수질 데이터 분석 시작...")
    root_results = analyze_water_quality(excel_file)

    print(root_results[0])

    # results = [well1, times, temps1, temps2, ecs1, ecs2, phs1, phs2]

    for i in range(1, num_sheets + 1):
        genhwpdoc.copy_template_to_desktop()
        results = root_results[i - 1]

        print('=' * 100)

        if len(results) == 5:  # 5 : single, 8 : dual
            well1 = results[0]
            times = results[1]
            temps1 = results[2]
            ecs1 = results[3]
            phs1 = results[4]

            ec1_max = ecs1[-3]
            ec1_min = ecs1[-2]

            temp1_max = temps1[-3]
            temp1_min = temps1[-2]

            ph1_max = phs1[-3]
            ph1_min = phs1[-2]
        else:
            well1 = results[0]
            well2 = "W-" + str(extract_number(well1) + 1)
            times = results[1]

            temps1 = results[2]
            temps2 = results[3]

            ecs1 = results[4]
            ecs2 = results[5]

            phs1 = results[6]
            phs2 = results[7]

            ec1_max = ecs1[-3]
            ec1_min = ecs1[-2]
            ec2_max = ecs2[-3]
            ec2_min = ecs2[-2]

            temp1_max = temps1[-3]
            temp1_min = temps1[-2]
            temp2_max = temps2[-3]
            temp2_min = temps2[-2]

            ph1_max = phs1[-3]
            ph1_min = phs1[-2]
            ph2_max = phs2[-3]
            ph2_min = phs2[-2]

        if len(results) == 5:
            print(results)
            print(well1)
            print(temps1)
            print(ecs1)
            print(phs1)
            hwp_part(results)
        else:
            print(results)
            print(well1)
            print(temps1)
            print(ecs1)
            print(phs1)
            print(well2)
            print(temps2)
            print(ecs2)
            print(phs2)
            hwp_part(results)

        print('=' * 100)

    merge_hwp_files()
    genhwpdoc.clean_up()


if __name__ == "__main__":
    main()
