"""
HWP Document Generator for Water Quality Analysis
Generates HWP documents from Excel water quality data.
"""

import os
import shutil
import datetime
from pathlib import Path
from typing import List, Tuple, Optional
from dataclasses import dataclass

import openpyxl
from pyhwpx import Hwp


@dataclass
class WaterQualityData:
    """Data class for water quality measurements."""
    well_name: str
    times: List[str]
    temperatures: List[float]
    ec_values: List[int]  # Electrical Conductivity
    ph_values: List[float]

    @property
    def temp_stats(self) -> Tuple[float, float]:
        """Returns (max, min) temperature from last 3 values."""
        return self.temperatures[-3], self.temperatures[-2]

    @property
    def ec_stats(self) -> Tuple[int, int]:
        """Returns (max, min) EC from last 3 values."""
        return self.ec_values[-3], self.ec_values[-2]

    @property
    def ph_stats(self) -> Tuple[float, float]:
        """Returns (max, min) pH from last 3 values."""
        return self.ph_values[-3], self.ph_values[-2]


class ExcelDataExtractor:
    """Extracts water quality data from Excel files."""

    # Cell coordinates for data extraction
    WELL_NAME_CELL = 'D12'
    TIME_COLUMN = 2  # Column B
    TEMP_COLUMN = 4  # Column D
    EC_COLUMN = 5  # Column E
    PH_COLUMN = 6  # Column F

    # Row ranges for data extraction
    TIMED_DATA_ROWS = range(14, 24)  # Rows with time stamps
    STATS_DATA_ROWS = range(24, 27)  # Rows with statistics only

    @staticmethod
    def extract_from_file(file_path: str) -> Optional[List[WaterQualityData]]:
        """
        Extract water quality data from all sheets in an Excel file.

        Args:
            file_path: Path to the Excel file

        Returns:
            List of WaterQualityData objects, one per sheet
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
        except FileNotFoundError:
            print(f"Error: File not found - {file_path}")
            return None
        except Exception as e:
            print(f"Error loading workbook: {e}")
            return None

        all_data = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            data = ExcelDataExtractor._extract_from_sheet(sheet)
            if data:
                all_data.append(data)

        return all_data

    @staticmethod
    def _extract_from_sheet(sheet) -> Optional[WaterQualityData]:
        """Extract water quality data from a single sheet."""
        well_name = sheet[ExcelDataExtractor.WELL_NAME_CELL].value
        print(f"Processing: {well_name or 'N/A'}")

        times, temps, ecs, phs = [], [], [], []

        # Extract timed data (rows with timestamps)
        for row in ExcelDataExtractor.TIMED_DATA_ROWS:
            time_val = sheet.cell(row=row, column=ExcelDataExtractor.TIME_COLUMN).value
            temp_val = sheet.cell(row=row, column=ExcelDataExtractor.TEMP_COLUMN).value
            ec_val = sheet.cell(row=row, column=ExcelDataExtractor.EC_COLUMN).value
            ph_val = sheet.cell(row=row, column=ExcelDataExtractor.PH_COLUMN).value

            if isinstance(time_val, datetime.datetime):
                times.append(time_val.strftime("%Y-%m-%d %H:%M"))

            if isinstance(temp_val, (int, float)):
                temps.append(round(temp_val, 1))

            if isinstance(ec_val, (int, float)):
                ecs.append(int(ec_val))

            if isinstance(ph_val, (int, float)):
                phs.append(round(ph_val, 2))

        # Extract statistics data (rows without timestamps)
        for row in ExcelDataExtractor.STATS_DATA_ROWS:
            temp_val = sheet.cell(row=row, column=ExcelDataExtractor.TEMP_COLUMN).value
            ec_val = sheet.cell(row=row, column=ExcelDataExtractor.EC_COLUMN).value
            ph_val = sheet.cell(row=row, column=ExcelDataExtractor.PH_COLUMN).value

            if isinstance(temp_val, (int, float)):
                temps.append(round(temp_val, 1))

            if isinstance(ec_val, (int, float)):
                ecs.append(int(ec_val))

            if isinstance(ph_val, (int, float)):
                phs.append(round(ph_val, 2))

        return WaterQualityData(
            well_name=well_name,
            times=times,
            temperatures=temps,
            ec_values=ecs,
            ph_values=phs
        )


class HwpDocumentGenerator:
    """Generates HWP documents from water quality data."""

    def __init__(self, template_path: Path, output_dir: Path, visible: bool = False):
        """
        Initialize the HWP document generator.

        Args:
            template_path: Path to the HWP template file
            output_dir: Directory for output files
            visible: Whether to show HWP application window
        """
        self.template_path = template_path
        self.output_dir = output_dir
        self.visible = visible
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate_document(self, data: WaterQualityData) -> Path:
        """
        Generate an HWP document from water quality data.

        Args:
            data: Water quality data to insert into template

        Returns:
            Path to the generated document
        """
        hwp = Hwp(visible=self.visible)
        hwp.open(str(self.template_path))

        try:
            self._fill_fields(hwp, data)
            self._fill_table(hwp, data)

            output_path = self.output_dir / f"ex_water_{data.well_name}.hwp"
            hwp.save_as(str(output_path))
            print(f"Generated: {output_path}")

            return output_path
        finally:
            hwp.Quit(save=False)

    def _fill_fields(self, hwp: Hwp, data: WaterQualityData):
        """Fill template fields with summary statistics."""
        temp_max, temp_min = data.temp_stats
        ec_max, ec_min = data.ec_stats
        ph_max, ph_min = data.ph_stats

        fields = {
            'well_main': data.well_name,
            'temp_min': temp_min,
            'temp_max': temp_max,
            'ec_min': ec_min,
            'ec_max': ec_max,
            'ph_min': ph_min,
            'ph_max': ph_max,
        }

        for field_name, value in fields.items():
            hwp.MoveToField(field_name)
            hwp.PutFieldText(field_name, str(value))

    def _fill_table(self, hwp: Hwp, data: WaterQualityData):
        """Fill the data table with measurements."""
        # Get and select the first table
        table_list = [ctrl for ctrl in hwp.ctrl_list if ctrl.UserDesc == "표"]
        if not table_list:
            raise ValueError("No table found in template")

        hwp.select_ctrl(table_list[0])
        hwp.ShapeObjTableSelCell()

        # Fill well name in header
        self._fill_cell(hwp, "C2", data.well_name)

        # Fill measurement data (rows 4-13)
        for i, row in enumerate(range(4, 14), start=0):
            self._fill_cell(hwp, f"A{row}", data.times[i])
            self._fill_cell(hwp, f"C{row}", str(data.temperatures[i]))
            self._fill_cell(hwp, f"D{row}", str(data.ec_values[i]))
            self._fill_cell(hwp, f"E{row}", str(data.ph_values[i]))

        # Fill statistics data (rows 14-16)
        for i, row in enumerate(range(14, 17), start=10):
            self._fill_cell(hwp, f"C{row}", str(data.temperatures[i]))
            self._fill_cell(hwp, f"D{row}", str(data.ec_values[i]))
            self._fill_cell(hwp, f"E{row}", str(data.ph_values[i]))

    @staticmethod
    def _fill_cell(hwp: Hwp, address: str, value: str):
        """Fill a single table cell with a value."""
        hwp.goto_addr(address)
        hwp.SelectAll()
        hwp.Delete()
        hwp.insert_text(value)


class TemplateManager:
    """Manages HWP template file operations."""

    def __init__(self, template_source: Path, working_dir: Path):
        """
        Initialize the template manager.

        Args:
            template_source: Source path of the template file
            working_dir: Working directory (e.g., Desktop)
        """
        self.template_source = template_source
        self.working_dir = working_dir
        self.template_name = template_source.name

    @property
    def working_template_path(self) -> Path:
        """Get the path to the template in the working directory."""
        return self.working_dir / self.template_name

    def copy_template(self) -> bool:
        """Copy template to working directory."""
        try:
            shutil.copy(self.template_source, self.working_dir)
            print(f"Template copied to: {self.working_template_path}")
            return True
        except Exception as e:
            print(f"Error copying template: {e}")
            return False

    def cleanup(self):
        """Remove template from working directory."""
        if self.working_template_path.exists():
            self.working_template_path.unlink()
            print(f"Cleaned up: {self.working_template_path}")


class WaterQualityReportGenerator:
    """Main application class for generating water quality reports."""

    def __init__(
            self,
            excel_file: Path,
            template_source: Path,
            output_dir: Path,
            visible: bool = False
    ):
        """
        Initialize the report generator.

        Args:
            excel_file: Path to Excel file with water quality data
            template_source: Path to HWP template file
            output_dir: Directory for output files
            visible: Whether to show HWP application window
        """
        self.excel_file = excel_file
        self.output_dir = output_dir
        self.visible = visible

        # Get desktop path for template operations
        desktop = Path(os.environ['USERPROFILE']) / 'Desktop'
        self.template_manager = TemplateManager(template_source, desktop)

    def generate_reports(self):
        """Generate HWP reports for all sheets in the Excel file."""
        # Validate Excel file
        if not self.excel_file.exists():
            print(f"Error: Excel file not found - {self.excel_file}")
            return

        # Extract data from Excel
        print("Extracting data from Excel...")
        all_data = ExcelDataExtractor.extract_from_file(str(self.excel_file))

        if not all_data:
            print("No data extracted from Excel file")
            return

        print(f"Found {len(all_data)} sheets to process\n")

        # Generate documents
        for data in all_data:
            print(f"\nProcessing well: {data.well_name}")

            # Copy template for this document
            if not self.template_manager.copy_template():
                print(f"Failed to copy template for {data.well_name}")
                continue

            # Generate document
            try:
                generator = HwpDocumentGenerator(
                    self.template_manager.working_template_path,
                    self.output_dir,
                    self.visible
                )
                generator.generate_document(data)
            except Exception as e:
                print(f"Error generating document for {data.well_name}: {e}")
            finally:
                # Clean up template
                self.template_manager.cleanup()

        print("\n✅ All reports generated successfully")


def main():
    """Main entry point."""
    # Configuration
    excel_file = Path("d:/05_Send/ex_water_test.xlsx")
    template_source = Path("c:/Program Files/totalcmd/hwp/wt_simple.hwp")
    output_dir = Path("d:/05_Send")

    # Create report generator
    generator = WaterQualityReportGenerator(
        excel_file=excel_file,
        template_source=template_source,
        output_dir=output_dir,
        visible=False  # Set to True for debugging
    )

    # Generate all reports
    generator.generate_reports()

    # Optionally merge HWP files (if merge_hwp_files function is available)
    try:
        from merge_hwp_files import merge_hwp_files
        merge_hwp_files()
    except ImportError:
        print("Note: merge_hwp_files module not found, skipping merge step")


if __name__ == "__main__":
    main()