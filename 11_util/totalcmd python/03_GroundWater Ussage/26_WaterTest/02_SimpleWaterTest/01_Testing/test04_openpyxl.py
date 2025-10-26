import openpyxl
import os
import datetime


def analyze_water_quality_openpyxl(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook.active

        well = sheet['D12'].value
        print(f"분석 대상: {well if well is not None else 'N/A'}")
        print("=" * 100)

        # 데이터 리스트 초기화
        times = []
        temps = []
        ecs = []
        phs = []

        print('='*100)
        for row in range(14, 24):  # 24는 포함되지 않으므로 14행부터 23행까지 반복
            # 셀 값 가져오기
            time_val = sheet.cell(row=row, column=2).value  # B열
            temp_val = sheet.cell(row=row, column=4).value  # D열
            ec_val = sheet.cell(row=row, column=5).value  # E열
            ph_val = sheet.cell(row=row, column=6).value  # F열

            # None이 아닌 경우에만 리스트에 추가 (win32com 코드와 동일한 로직)
            # 수치형 데이터인지 확인 (숫자가 아닌 문자열이나 None은 제외)

            if isinstance(time_val, datetime.datetime):
                times.append(time_val)
                print(time_val)

            if isinstance(temp_val, (int, float)):
                temps.append(round(temp_val,1))

            if isinstance(ec_val, (int, float)):
                ecs.append(int(ec_val))

            if isinstance(ph_val, (int, float)):
                phs.append(round(ph_val,2))


        for row in range(24, 27):  # 27는 포함되지 않으므로 24행부터 26행까지 반복
            # 셀 값 가져오기
            temp_val = sheet.cell(row=row, column=4).value  # D열
            ec_val = sheet.cell(row=row, column=5).value  # E열
            ph_val = sheet.cell(row=row, column=6).value  # F열

            if isinstance(temp_val, (int, float)):
                temps.append(round(temp_val,1))

            if isinstance(ec_val, (int, float)):
                ecs.append(int(ec_val))

            if isinstance(ph_val, (int, float)):
                phs.append(round(ph_val,2))

        results = [times, temps, ecs, phs]

        return well, results

    except FileNotFoundError:
        print(f"오류 발생: 파일을 찾을 수 없습니다 - {file_path}")
        return None
    except KeyError:
        print("오류 발생: 워크시트를 찾을 수 없습니다.")
        return None
    except Exception as e:
        print(f"오류 발생: {e}")
        return None


if __name__ == "__main__":
    file_path = r"ex_water_test.xlsx"

    if not os.path.exists(file_path):
        print(f"파일을 찾을 수 없습니다: {file_path}")
        print("파일 경로를 확인해주세요. (openpyxl은 .xlsx 파일만 지원합니다.)")
    else:
        # 분석 실행
        print("📢 openpyxl을 사용하여 수질 데이터 분석 시작...")
        well, results = analyze_water_quality_openpyxl(file_path)

        times = results[0]
        temps = results[1]
        ecs = results[2]
        phs = results[3]

        print(temps)
        print(ecs)
        print(phs)

        ec_max = ecs[-3]
        ec_min = ecs[-2]

        temp_max = temps[-3]
        temp_min = temps[-2]

        ph_max = phs[-3]
        ph_min = phs[-2]

        print(well)
        print(ph_max, ph_min)
        print(temp_max, temp_min)
        print(ec_max, ec_min)


        if results:
            print("=" * 100)
            print("분석 완료!")