import re
import os
import win32com.client as win32
from natsort import natsorted

from pyhwpx import Hwp

import pandas as pd
import openpyxl
import random
import pyautogui
import pygetwindow as gw
import time

from datetime import datetime, timedelta
from typing import Any

import tkinter as tk

import openpyxl
import os
import datetime


def analyze_water_quality_openpyxl(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook.active

        well = sheet['D12'].value
        print(f"분석 대상: {well if well is not None else 'N/A'}")

        # 데이터 리스트 초기화
        times, temps, ecs, phs = [], [], [], []

        print('=' * 100)
        for row in range(14, 24):  # 24는 포함되지 않으므로 14행부터 23행까지 반복
            # 셀 값 가져오기
            time_val = sheet.cell(row=row, column=2).value  # B열
            temp_val = sheet.cell(row=row, column=4).value  # D열
            ec_val = sheet.cell(row=row, column=5).value  # E열
            ph_val = sheet.cell(row=row, column=6).value  # F열

            # None이 아닌 경우에만 리스트에 추가 (win32com 코드와 동일한 로직)
            # 수치형 데이터인지 확인 (숫자가 아닌 문자열이나 None은 제외)

            if isinstance(time_val, datetime.datetime):
                times.append(time_val.strftime("%Y-%m-%d %H:%M"))
                print(time_val.strftime("%Y-%m-%d %H:%M"))

            if isinstance(temp_val, (int, float)):
                temps.append(round(temp_val, 1))

            if isinstance(ec_val, (int, float)):
                ecs.append(int(ec_val))

            if isinstance(ph_val, (int, float)):
                phs.append(round(ph_val, 2))

        for row in range(24, 27):  # 27는 포함되지 않으므로 24행부터 26행까지 반복
            # 셀 값 가져오기
            temp_val = sheet.cell(row=row, column=4).value  # D열
            ec_val = sheet.cell(row=row, column=5).value  # E열
            ph_val = sheet.cell(row=row, column=6).value  # F열

            if isinstance(temp_val, (int, float)):
                temps.append(round(temp_val, 1))

            if isinstance(ec_val, (int, float)):
                ecs.append(int(ec_val))

            if isinstance(ph_val, (int, float)):
                phs.append(round(ph_val, 2))

        results = [times, temps, ecs, phs]

        return well, results

    except FileNotFoundError:
        print(f"오류 발생: 파일을 찾을 수 없습니다 - {file_path}")
        return None
    except KeyError:
        print("오류 발생: 워크시트를 찾을 수 없습니다.")
        return None
    except Exception as e:
        print(f"오류 발생: {e}")
        return None


def main():
    file_path = r"d:\05_Send\pythonProject\03_GroundWater Ussage\26_WaterTest\02_SimpleWaterTest\ex_water_test.xlsx"

    if not os.path.exists(file_path):
        print(f"파일을 찾을 수 없습니다: {file_path}")
        print("파일 경로를 확인해주세요. (openpyxl은 .xlsx 파일만 지원합니다.)")
    else:
        print("📢 openpyxl을 사용하여 수질 데이터 분석 시작...")
        well, results = analyze_water_quality_openpyxl(file_path)

    times = results[0]
    temps = results[1]
    ecs = results[2]
    phs = results[3]

    print(temps)
    print(ecs)
    print(phs)

    ec_max = ecs[-3]
    ec_min = ecs[-2]

    temp_max = temps[-3]
    temp_min = temps[-2]

    ph_max = phs[-3]
    ph_min = phs[-2]

    hwp = Hwp()

    hwp.MoveToField('well_main')
    hwp.PutFieldText('well_main', f'{well}')

    hwp.MoveToField('temp_min')
    hwp.PutFieldText('temp_min', f'{temp_min}')

    hwp.MoveToField('temp_max')
    hwp.PutFieldText('temp_max', f'{temp_max}')

    hwp.MoveToField('ec_min')
    hwp.PutFieldText('ec_min', f'{ec_min}')

    hwp.MoveToField('ec_max')
    hwp.PutFieldText('ec_max', f'{ec_max}')

    hwp.MoveToField('ph_min')
    hwp.PutFieldText('ph_min', f'{ph_min}')

    hwp.MoveToField('ph_max')
    hwp.PutFieldText('ph_max', f'{ph_max}')



    # 테이블리스트를 가져옴
    Table_list = [ i for i in hwp.ctrl_list if i.UserDesc =="표"]
    print(Table_list)

    #테이블의 첫번째를 선택
    hwp.select_ctrl(Table_list[0])
    hwp.ShapeObjTableSelCell()  # A1 셀을 선택한 상태가 됨


    hwp.goto_addr("C2")
    hwp.SelectAll()
    hwp.Delete()
    hwp.insert_text(f"{well}")

    print(well)
    print(ph_max, ph_min)
    print(temp_max, temp_min)
    print(ec_max, ec_min)

    for i in range(4,14):
        hwp.goto_addr(f"A{i}")
        hwp.SelectAll()
        hwp.Delete()
        hwp.insert_text(f"{times[i-4]}")

        hwp.goto_addr(f"C{i}")
        hwp.SelectAll()
        hwp.Delete()
        hwp.insert_text(f"{temps[i-4]}")

        hwp.goto_addr(f"D{i}")
        hwp.SelectAll()
        hwp.Delete()
        hwp.insert_text(f"{ecs[i - 4]}")

        hwp.goto_addr(f"E{i}")
        hwp.SelectAll()
        hwp.Delete()
        hwp.insert_text(f"{phs[i - 4]}")



    for i in range(14,17):
        hwp.goto_addr(f"C{i}")
        hwp.SelectAll()
        hwp.Delete()
        hwp.insert_text(f"{temps[i-4]}")

        hwp.goto_addr(f"D{i}")
        hwp.SelectAll()
        hwp.Delete()
        hwp.insert_text(f"{ecs[i - 4]}")

        hwp.goto_addr(f"E{i}")
        hwp.SelectAll()
        hwp.Delete()
        hwp.insert_text(f"{phs[i - 4]}")



def TableFindSample():
    # Table 찾아가기
    # https://www.inflearn.com/community/questions/1453780/%ED%91%9C-%EC%B0%BE%EC%95%84%EA%B0%80%EA%B8%B0-%EB%8F%84%EC%99%80%EC%A3%BC%EC%84%B8%EC%9A%94-%E3%85%A0%E3%85%A0

    Table_list = [ i for i in hwp.ctrl_list if i.UserDesc =="표"]
    print(Table_list)

    hwp.select_ctrl(Table_list[0])
    hwp.ShapeObjTableSelCell()
    hwp.goto_addr("C2")
    hwp.insert_text(f"{well}")

    #=========================================================================

    hwp.get_into_nth_table(0)
    hwp.goto_addr("C2")
    hwp.insert_text(f"{well}")

    #=========================================================================

    Table_list = [i for i in hwp.ctrl_list if i.UserDesc == "표"]
    Table_index = 0

    hwp.move_to_ctrl(Table_list[Table_index])
    hwp.find_ctrl()
    hwp.ShapeObjTableSelCell()  # A1 셀을 선택한 상태가 됨
    hwp.goto_addr("C2")
    hwp.SelectAll()
    hwp.Delete()
    hwp.insert_text(f"{well}")
    #=========================================================================
    hwp.ShapeObjTableSelCell()
    hwp.TableCellBlock()
    hwp.TableCellBlockExtend()
    hwp.TableCellBlockExtend()





if __name__ == "__main__":
    main()
