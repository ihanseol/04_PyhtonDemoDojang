import re
from pyhwpx import Hwp
import openpyxl
from pathlib import Path
import os
import shutil
import time
import datetime
from merge_hwp_files import merge_hwp_files


class GenHwpDoc:
    """Class to generate HWP documents from Excel data."""

    def __init__(self):
        """Initialize the document generator with default paths."""
        self.base_dir = Path("d:/05_Send")
        self.excel_file = "ex_water_test.xlsx"
        self.hwp_template = "wt_simple.hwp"
        self.template_source = Path("c:/Program Files/totalcmd/hwp") / self.hwp_template
        self.desktop = self.get_desktop_path()

    @staticmethod
    def get_desktop_path():
        """Get the path to the user's desktop."""
        return Path(os.environ['USERPROFILE']) / 'Desktop'

    def get_hwp_name(self):
        return self.desktop / self.hwp_template


    def clean_up(self):
        template_path = self.desktop / self.hwp_template
        if template_path.exists():
            os.remove(template_path)


    def copy_template_to_desktop(self):
        """Copy the HWP template file to the desktop."""
        try:
            desktop_template = self.desktop / self.hwp_template
            shutil.copy(self.template_source, self.desktop)
            return True
        except Exception as e:
            print(f"Error copying template to desktop: {str(e)}")
            return False



def analyze_water_quality_openpyxl(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)

    except FileNotFoundError:
        print(f"오류 발생: 파일을 찾을 수 없습니다 - {file_path}")
        return None
    except Exception as e:
        print(f"오류 발생: {e}")
        return None

    return_data = []
    sheet_names = workbook.sheetnames

    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        well = sheet['D12'].value
        print(f"분석 대상: {well if well is not None else 'N/A'}")

        # 데이터 리스트 초기화
        times, temps, ecs, phs = [], [], [], []

        print('=' * 100)
        for row in range(14, 24):  # 24는 포함되지 않으므로 14행부터 23행까지 반복
            # 셀 값 가져오기
            time_val = sheet.cell(row=row, column=2).value  # B열
            temp_val = sheet.cell(row=row, column=4).value  # D열
            ec_val = sheet.cell(row=row, column=5).value  # E열
            ph_val = sheet.cell(row=row, column=6).value  # F열

            # None이 아닌 경우에만 리스트에 추가 (win32com 코드와 동일한 로직)
            # 수치형 데이터인지 확인 (숫자가 아닌 문자열이나 None은 제외)

            if isinstance(time_val, datetime.datetime):
                times.append(time_val.strftime("%Y-%m-%d %H:%M"))
                print(time_val.strftime("%Y-%m-%d %H:%M"))

            if isinstance(temp_val, (int, float)):
                temps.append(round(temp_val, 1))

            if isinstance(ec_val, (int, float)):
                ecs.append(int(ec_val))

            if isinstance(ph_val, (int, float)):
                phs.append(round(ph_val, 2))

        for row in range(24, 27):  # 27는 포함되지 않으므로 24행부터 26행까지 반복
            # 셀 값 가져오기
            temp_val = sheet.cell(row=row, column=4).value  # D열
            ec_val = sheet.cell(row=row, column=5).value  # E열
            ph_val = sheet.cell(row=row, column=6).value  # F열

            if isinstance(temp_val, (int, float)):
                temps.append(round(temp_val, 1))

            if isinstance(ec_val, (int, float)):
                ecs.append(int(ec_val))

            if isinstance(ph_val, (int, float)):
                phs.append(round(ph_val, 2))

        results = [well, times, temps, ecs, phs]
        return_data.append(results)

    return return_data


def hwp_part(results):
    # hwp = Hwp(visible=True)
    hwp = Hwp(visible=False)
    hwp.open(r"c:\Users\minhwasoo\Desktop\wt_simple.hwp")

    well = results[0]
    times = results[1]
    temps = results[2]
    ecs = results[3]
    phs = results[4]

    ec_max = ecs[-3]
    ec_min = ecs[-2]

    temp_max = temps[-3]
    temp_min = temps[-2]

    ph_max = phs[-3]
    ph_min = phs[-2]

    hwp.MoveToField('well_main')
    hwp.PutFieldText('well_main', f'{well}')

    hwp.MoveToField('temp_min')
    hwp.PutFieldText('temp_min', f'{temp_min}')

    hwp.MoveToField('temp_max')
    hwp.PutFieldText('temp_max', f'{temp_max}')

    hwp.MoveToField('ec_min')
    hwp.PutFieldText('ec_min', f'{ec_min}')

    hwp.MoveToField('ec_max')
    hwp.PutFieldText('ec_max', f'{ec_max}')

    hwp.MoveToField('ph_min')
    hwp.PutFieldText('ph_min', f'{ph_min}')

    hwp.MoveToField('ph_max')
    hwp.PutFieldText('ph_max', f'{ph_max}')

    # 테이블리스트를 가져옴
    Table_list = [i for i in hwp.ctrl_list if i.UserDesc == "표"]
    print(Table_list)

    # 테이블의 첫번째를 선택
    hwp.select_ctrl(Table_list[0])
    hwp.ShapeObjTableSelCell()  # A1 셀을 선택한 상태가 됨

    hwp.goto_addr("C2")
    hwp.SelectAll()
    hwp.Delete()
    hwp.insert_text(f"{well}")

    print(well)
    print(ph_max, ph_min)
    print(temp_max, temp_min)
    print(ec_max, ec_min)

    for i in range(4, 14):
        hwp.goto_addr(f"A{i}")
        hwp.SelectAll()
        hwp.Delete()
        hwp.insert_text(f"{times[i - 4]}")

        hwp.goto_addr(f"C{i}")
        hwp.SelectAll()
        hwp.Delete()
        hwp.insert_text(f"{temps[i - 4]}")

        hwp.goto_addr(f"D{i}")
        hwp.SelectAll()
        hwp.Delete()
        hwp.insert_text(f"{ecs[i - 4]}")

        hwp.goto_addr(f"E{i}")
        hwp.SelectAll()
        hwp.Delete()
        hwp.insert_text(f"{phs[i - 4]}")

    for i in range(14, 17):
        hwp.goto_addr(f"C{i}")
        hwp.SelectAll()
        hwp.Delete()
        hwp.insert_text(f"{temps[i - 4]}")

        hwp.goto_addr(f"D{i}")
        hwp.SelectAll()
        hwp.Delete()
        hwp.insert_text(f"{ecs[i - 4]}")

        hwp.goto_addr(f"E{i}")
        hwp.SelectAll()
        hwp.Delete()
        hwp.insert_text(f"{phs[i - 4]}")

    hwp.save_as(f"d:\\05_Send\\ex_water_{well}.hwp")
    hwp.Quit(save=False)

def copy_template_to_desktop(self):
    try:
        desktop_template = desktop / hwp_template
        shutil.copy(self.template_source, desktop_template)
        return True
    except Exception as e:
        print(f"Error copying template to desktop: {str(e)}")
        return False


def main():
    excel_file = r"d:\05_Send\ex_water_test.xlsx"
    genhwpdoc = GenHwpDoc()

    wb = openpyxl.load_workbook(excel_file)

    # 모든 시트 이름 가져오기
    sheet_names = wb.sheetnames
    num_sheets = len(sheet_names)

    print(f"발견된 시트: {sheet_names}\n")

    if not os.path.exists(excel_file):
        print(f"파일을 찾을 수 없습니다: {excel_file}")
        print("파일 경로를 확인해주세요. (openpyxl은 .xlsx 파일만 지원합니다.)")
        exit(0)

    print("📢 openpyxl을 사용하여 수질 데이터 분석 시작...")
    root_results = analyze_water_quality_openpyxl(excel_file)

    print(root_results)


    for i in range(1, num_sheets + 1):
        genhwpdoc.copy_template_to_desktop()

        results = root_results[i - 1]
        print('=' * 100)

        well = results[0]
        times = results[1]
        temps = results[2]
        ecs = results[3]
        phs = results[4]

        ec_max = ecs[-3]
        ec_min = ecs[-2]

        temp_max = temps[-3]
        temp_min = temps[-2]

        ph_max = phs[-3]
        ph_min = phs[-2]

        print(results)
        print(well)
        print(temps)
        print(ecs)
        print(phs)

        print('=' * 100)

        hwp_part(results)

    genhwpdoc.clean_up()
    merge_hwp_files()



if __name__ == "__main__":
    main()
