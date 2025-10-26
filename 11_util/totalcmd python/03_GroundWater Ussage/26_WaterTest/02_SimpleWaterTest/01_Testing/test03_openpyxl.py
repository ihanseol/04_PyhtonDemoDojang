import openpyxl
import os


def analyze_water_quality_openpyxl(file_path):
    try:
        # 워크북 로드
        # data_only=True 옵션을 사용하여 셀의 수식 대신 계산된 값을 읽어옵니다.
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        # 활성화된 시트 (보통 첫 번째 시트) 선택
        sheet = workbook.active

        # Well 정보 읽기 (D12 셀)
        # openpyxl은 1-based 인덱스를 사용하거나, Excel의 셀 주소를 직접 사용합니다.
        well = sheet['D12'].value

        # None 값을 처리하여 문자열로 출력
        print(f"분석 대상: {well if well is not None else 'N/A'}")
        print("=" * 50)

        # 데이터 리스트 초기화
        times = []
        temps = []
        ecs = []
        phs = []

        # 데이터 읽기 (14행부터 23행까지)
        # openpyxl은 셀 주소 대신 행/열 인덱스를 사용하여 셀에 접근할 수도 있습니다.
        # 행: 14 ~ 23 (Excel 행 번호)
        # 열: B(2), D(4), E(5), F(6)
        for row in range(14, 24):  # 24는 포함되지 않으므로 14행부터 23행까지 반복
            # 셀 값 가져오기
            time_val = sheet.cell(row=row, column=2).value  # B열
            temp_val = sheet.cell(row=row, column=4).value  # D열
            ec_val = sheet.cell(row=row, column=5).value  # E열
            ph_val = sheet.cell(row=row, column=6).value  # F열

            # None이 아닌 경우에만 리스트에 추가 (win32com 코드와 동일한 로직)
            # 수치형 데이터인지 확인 (숫자가 아닌 문자열이나 None은 제외)
            if isinstance(time_val, (int, float, str, type(None))):  # openpyxl은 time_val이 날짜/시간 객체일 수 있음
                times.append(time_val)

            if isinstance(temp_val, (int, float)):
                temps.append(temp_val)

            if isinstance(ec_val, (int, float)):
                ecs.append(ec_val)

            if isinstance(ph_val, (int, float)):
                phs.append(ph_val)

        # 결과 딕셔너리 초기화
        results = {
            'well': well if well is not None else None,
            'temp_max': None, 'temp_min': None,
            'ec_max': None, 'ec_min': None,
            'ph_max': None, 'ph_min': None
        }

        # 최대/최소값 계산 및 출력
        if temps:
            temp_max = round(max(temps), 1)
            temp_min = round(min(temps), 1)
            print(f"온도(°C)")
            print(f"  최대값: {temp_max}")
            print(f"  최소값: {temp_min}")
            print()
            results['temp_max'] = temp_max
            results['temp_min'] = temp_min

        if ecs:
            # 원본 코드와 같이 반올림 후 정수형으로 변환
            ec_max = int(round(max(ecs)))
            ec_min = int(round(min(ecs)))
            print(f"EC(μS/cm)")
            print(f"  최대값: {ec_max}")
            print(f"  최소값: {ec_min}")
            print()
            results['ec_max'] = ec_max
            results['ec_min'] = ec_min

        if phs:
            ph_max = round(max(phs), 2)
            ph_min = round(min(phs), 2)
            print(f"pH")
            print(f"  최대값: {ph_max}")
            print(f"  최소값: {ph_min}")
            print()
            results['ph_max'] = ph_max
            results['ph_min'] = ph_min

        return results

    except FileNotFoundError:
        print(f"오류 발생: 파일을 찾을 수 없습니다 - {file_path}")
        return None
    except KeyError:
        print("오류 발생: 워크시트를 찾을 수 없습니다.")
        return None
    except Exception as e:
        print(f"오류 발생: {e}")
        return None
    # openpyxl은 win32com처럼 명시적인 workbook.Close()나 excel.Quit()가 필요하지 않습니다.


if __name__ == "__main__":
    # Excel 파일 경로 설정 (본인의 파일 경로로 수정하세요)
    # **주의: openpyxl을 사용하려면 파일이 .xlsx 형식이어야 합니다.**
    file_path = r"ex_water_test.xlsx"

    # 파일 존재 확인
    if not os.path.exists(file_path):
        print(f"파일을 찾을 수 없습니다: {file_path}")
        print("파일 경로를 확인해주세요. (openpyxl은 .xlsx 파일만 지원합니다.)")
    else:
        # 분석 실행
        print("📢 openpyxl을 사용하여 수질 데이터 분석 시작...")
        results = analyze_water_quality_openpyxl(file_path)

        if results:
            print("=" * 50)
            print("분석 완료!")